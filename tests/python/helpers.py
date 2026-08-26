"""Monta workbooks openpyxl mínimos em memória pros testes -- evita depender
da planilha real (K19-204-...xlsx) pra testar lógica de leitura/agregação."""
import openpyxl


def build_workbook(sheets):
    """sheets: {nome_da_aba: [(linha, coluna, valor), ...]} (1-based, como openpyxl)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cells in sheets.items():
        ws = wb.create_sheet(name)
        for row, col, value in cells:
            ws.cell(row=row, column=col, value=value)
    return wb
