# -*- coding: utf-8 -*-
"""
Extrai os dados da planilha MSI (Controle, listas de documentos e as
não conformidades registradas em RIs) para um JSON que alimenta o
dashboard HTML.

Reexecute este script sempre que a planilha for atualizada, e depois
regenere o dashboard (build_dashboard.py) a partir do JSON gerado.

Nota: a partir da revisão R05 a planilha não tem mais RDO/OS/Obsoletos
(o usuário removeu essas abas), e o Controle foi reorganizado (cabeçalho
começa na linha 2, dados a partir da linha 4).
"""
import json
import sys
import datetime
import openpyxl

CONTRATO_INFO = {
    "contrato": "089/2026",
    "cliente": "Ferro+ Mineração S.A",
    "empresa": "MSI Engenharia e Tecnologia",
    "objeto": "Sistema de Proteção contra Descargas Atmosféricas (SPDA)",
}

DOC_SHEETS = {
    "Desenhos (DE)": ("Desenhos", 5, 92),
    "Memoriais Descritivos (MD)": ("MDs", 4, 87),
    "Listas de Materiais (LM)": ("LMs", 4, 89),
    "Análises de Risco (MC)": ("MCs", 4, 98),
    "Relatórios de Inspeção (RI)": ("RIs", 4, 99),
}

RI_DATA_ROW = 4
NC_DESC_COL, NC_SEV_COL, NC_STATUS_COL, NC_RESP_COL, NC_DATE_COL = 13, 14, 15, 16, 17
SEVERIDADE_ORDER = {"Crítica": 0, "Média": 1, "Baixa": 2, "Não informado": 3}


def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def last_row_with_value(ws, col, floor, ceiling=None):
    """Dynamic last-row detection so the script keeps working as rows are added."""
    ceiling = ceiling or ws.max_row
    last = floor - 1
    for r in range(floor, ceiling + 1):
        if ws.cell(row=r, column=col).value:
            last = r
    return max(last, floor - 1)


# ---------------------------------------------------------------------------
# Leitura (planilha -> listas de dicts). Nenhuma agregação acontece aqui --
# cada função só lê o que está na aba, já convertendo tipos (datas -> ISO,
# texto -> str().strip()).
# ---------------------------------------------------------------------------
def read_areas(wb):
    ws = wb["Controle"]
    data_row = 4
    last_row = last_row_with_value(ws, 5, data_row, 200)  # col E = descrição

    areas = []
    for r in range(data_row, last_row + 1):
        desc = ws.cell(row=r, column=5).value  # E: Descricao da Area
        if not desc:
            continue
        validade_laudo = as_date(ws.cell(row=r, column=20).value)  # T
        validade_is = as_date(ws.cell(row=r, column=23).value)  # W
        pendencia = ws.cell(row=r, column=29).value  # AC
        areas.append({
            "row": r,
            "codigo_ld": ws.cell(row=r, column=2).value,  # B
            "descricao": str(desc).strip(),
            "status": ws.cell(row=r, column=26).value,  # Z
            "adequacao_geral": ws.cell(row=r, column=16).value,  # P
            "validade_laudo": validade_laudo.isoformat() if validade_laudo else None,
            "validade_is": validade_is.isoformat() if validade_is else None,
            "pendencia": (str(pendencia).strip() if pendencia else None),
            "dossie": ws.cell(row=r, column=27).value,  # AA
        })
    return areas


def read_documentacao_counts(wb):
    counts = {}
    for label, (sheet, drow, floor_last) in DOC_SHEETS.items():
        sws = wb[sheet]
        last = last_row_with_value(sws, 1, drow, max(floor_last, sws.max_row))
        counts[label] = sum(1 for r in range(drow, last + 1) if sws.cell(row=r, column=1).value)
    return counts


def read_nao_conformidades(wb):
    """Aba RIs, colunas M/N/O/P/Q = Descrição/Severidade/Status/Responsável/Data."""
    ws = wb["RIs"]
    last_row = last_row_with_value(ws, 1, RI_DATA_ROW, max(99, ws.max_row))

    rows = []
    for r in range(RI_DATA_ROW, last_row + 1):
        desc = ws.cell(row=r, column=NC_DESC_COL).value
        if not desc or not str(desc).strip():
            continue
        area = ws.cell(row=r, column=3).value  # Descrição da Área
        data_nc = as_date(ws.cell(row=r, column=NC_DATE_COL).value)
        rows.append({
            "area": str(area).strip() if area else None,
            "numero_ri": ws.cell(row=r, column=1).value,  # Número (K19-204-...-RI-xxx)
            "descricao": str(desc).strip(),
            "severidade": ws.cell(row=r, column=NC_SEV_COL).value or "Não informado",
            "status": ws.cell(row=r, column=NC_STATUS_COL).value or "Não informado",
            "responsavel": ws.cell(row=r, column=NC_RESP_COL).value,
            "data": data_nc.isoformat() if data_nc else None,
        })
    return rows


# ---------------------------------------------------------------------------
# Agregação (listas de dicts -> métricas). Lógica pura, sem openpyxl --
# testável sem precisar montar uma planilha.
# ---------------------------------------------------------------------------
def compute_status_counts(areas):
    counts = {}
    for a in areas:
        s = a["status"] or "Não informado"
        counts[s] = counts.get(s, 0) + 1
    return counts


def compute_adequacao_media(areas):
    valores = [a["adequacao_geral"] for a in areas if isinstance(a["adequacao_geral"], (int, float))]
    return sum(valores) / len(valores) if valores else None


def compute_pendencias_abertas(areas):
    return [
        {"area": a["descricao"], "codigo": a["codigo_ld"], "pendencia": a["pendencia"]}
        for a in areas
        if a["pendencia"] and a["pendencia"].upper() not in ("OK", "-", "")
    ]


def compute_vencimentos(areas, today):
    def days_diff(iso_date):
        d = datetime.date.fromisoformat(iso_date)
        return (d - today).days

    vencimentos = []
    for a in areas:
        for campo, label in (("validade_laudo", "Laudo completo (medição)"), ("validade_is", "Inspeção Semestral (IS)")):
            v = a[campo]
            if not v:
                continue
            dd = days_diff(v)
            if dd <= 90:  # vencido ou vencendo nos proximos 90 dias
                vencimentos.append({
                    "area": a["descricao"],
                    "codigo": a["codigo_ld"],
                    "documento": label,
                    "data": v,
                    "dias": dd,
                    "situacao": "vencido" if dd < 0 else "vencendo",
                })
    vencimentos.sort(key=lambda x: x["dias"])
    return vencimentos


def sort_nao_conformidades(rows):
    return sorted(rows, key=lambda x: (SEVERIDADE_ORDER.get(x["severidade"], 9),
                                        0 if x["status"] == "Aberta" else 1))


def compute_nc_counts(rows):
    abertas = sum(1 for n in rows if n["status"] == "Aberta")
    status_counts, severidade_counts = {}, {}
    for n in rows:
        status_counts[n["status"]] = status_counts.get(n["status"], 0) + 1
        severidade_counts[n["severidade"]] = severidade_counts.get(n["severidade"], 0) + 1
    return abertas, status_counts, severidade_counts


# ---------------------------------------------------------------------------
# Orquestração: monta o dicionário final que vira dashboard_data.json.
# ---------------------------------------------------------------------------
def build_dashboard_data(wb, today):
    areas = read_areas(wb)
    nao_conformidades = sort_nao_conformidades(read_nao_conformidades(wb))
    nc_abertas, nc_status_counts, nc_severidade_counts = compute_nc_counts(nao_conformidades)

    return {
        "gerado_em": today.isoformat(),
        **CONTRATO_INFO,
        "areas_total": len(areas),
        "status_counts": compute_status_counts(areas),
        "adequacao_media": compute_adequacao_media(areas),
        "documentacao": read_documentacao_counts(wb),
        "pendencias_abertas": compute_pendencias_abertas(areas),
        "vencimentos": compute_vencimentos(areas, today),
        "nao_conformidades": nao_conformidades,
        "nc_abertas": nc_abertas,
        "nc_status_counts": nc_status_counts,
        "nc_severidade_counts": nc_severidade_counts,
    }


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "K19-204-FER-LD-001-R05 - Lista de Documentos (MSI).xlsx"
    today = datetime.date.fromisoformat(sys.argv[2]) if len(sys.argv) > 2 else datetime.date.today()
    out_path = sys.argv[3] if len(sys.argv) > 3 else "dashboard_data.json"

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = build_dashboard_data(wb, today)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("areas_total:", data["areas_total"])
    print("status_counts:", data["status_counts"])
    print("adequacao_media:", data["adequacao_media"])
    print("documentacao:", data["documentacao"])
    print("pendencias_abertas:", len(data["pendencias_abertas"]))
    print("vencimentos (<=90d):", len(data["vencimentos"]))
    print("nao_conformidades:", len(data["nao_conformidades"]), "abertas:", data["nc_abertas"])
    print("Saved ->", out_path)


if __name__ == "__main__":
    main()
