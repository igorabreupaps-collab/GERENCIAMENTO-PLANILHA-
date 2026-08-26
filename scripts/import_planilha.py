# -*- coding: utf-8 -*-
"""
Importa TODOS os dados da planilha oficial (.xlsx) para o Postgres, através
da própria API REST do painel (não fala direto com o banco -- passa pelas
mesmas validações e regras que a interface usa).

Uso:
    python3 scripts/import_planilha.py "K19-204-FER-LD-001-R05 - Lista de Documentos (MSI).xlsx"

Pré-requisitos:
    - O stack precisa estar no ar (docker compose up) com o schema já
      aplicado e pelo menos um usuário admin/editor existente (o admin de
      bootstrap criado automaticamente serve).
    - Variáveis de ambiente (ou um .env na raiz do projeto):
        API_BASE_URL          -- default: http://localhost:3000
        IMPORT_ADMIN_EMAIL    -- e-mail de um usuário admin ou editor
        IMPORT_ADMIN_PASSWORD

O que é importado:
    - Áreas (aba Controle) -> tabela areas
    - Não conformidades (seção da aba RIs) -> tabela nao_conformidades,
      vinculadas à área por nome (quando bate exatamente)
    - Documentos (abas Desenhos/MDs/LMs/MCs/RIs) -> tabela documentos,
      também vinculados à área por nome. Os múltiplos campos "Título 1..6"
      inconsistentes entre abas da planilha original são consolidados num
      único campo "titulo" por documento.

Idempotência: a API sempre insere linha nova (não existe "upsert" por
número/nome) -- rodar este script duas vezes duplica os dados. Ele recusa
rodar se a tabela "areas" já tiver alguma linha, a menos que --force seja
passado conscientemente.
"""
import datetime
import os
import sys

import openpyxl
import requests

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def load_dotenv(path):
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


class ConfigError(Exception):
    """Config obrigatória ausente/inválida -- main() decide como reportar."""


def resolve_config(argv, env):
    """Lê CLI args + variáveis de ambiente (sem tocar o processo: não dá
    load_dotenv nem sys.exit aqui) e devolve um dict com tudo que o import
    precisa. Extraído do nível de módulo pra ser testável sem precisar de
    variáveis de ambiente reais nem de um .env no disco."""
    admin_email = env.get("IMPORT_ADMIN_EMAIL") or env.get("BOOTSTRAP_ADMIN_EMAIL")
    admin_password = env.get("IMPORT_ADMIN_PASSWORD") or env.get("BOOTSTRAP_ADMIN_PASSWORD")
    if not admin_email or not admin_password:
        raise ConfigError(
            "Faltam as credenciais de quem vai importar.\n"
            "Defina IMPORT_ADMIN_EMAIL e IMPORT_ADMIN_PASSWORD no seu .env\n"
            "(pode reaproveitar BOOTSTRAP_ADMIN_EMAIL/BOOTSTRAP_ADMIN_PASSWORD, se for o mesmo admin)."
        )

    xlsx_arg = argv[1] if len(argv) > 1 and not argv[1].startswith("--") else None
    return {
        "api_base_url": env.get("API_BASE_URL", "http://localhost:3000").rstrip("/"),
        "admin_email": admin_email,
        "admin_password": admin_password,
        "xlsx_path": xlsx_arg or "K19-204-FER-LD-001-R05 - Lista de Documentos (MSI).xlsx",
        "force": "--force" in argv,
    }


def login(session, api_base_url, email, password):
    r = session.post(f"{api_base_url}/api/auth/login", json={"email": email, "password": password})
    if not r.ok:
        sys.exit(f"Não foi possível logar em {api_base_url}: {r.status_code} {r.text}")
    session.headers.update({"Authorization": f"Bearer {r.json()['token']}"})


def api_get(session, api_base_url, path):
    r = session.get(f"{api_base_url}{path}")
    r.raise_for_status()
    return r.json()


def api_post(session, api_base_url, path, body):
    r = session.post(f"{api_base_url}{path}", json=body)
    if not r.ok:
        raise RuntimeError(f"POST {path} falhou ({r.status_code}): {r.text}")
    return r.json()


# ---------------------------------------------------------------------------
# Leitura do xlsx
# ---------------------------------------------------------------------------
def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def last_row_with_value(ws, col, floor, ceiling=None):
    ceiling = ceiling or ws.max_row
    last = floor - 1
    for r in range(floor, ceiling + 1):
        if ws.cell(row=r, column=col).value:
            last = r
    return max(last, floor - 1)


def read_areas(wb):
    ws = wb["Controle"]
    data_row = 4
    last_row = last_row_with_value(ws, 5, data_row, max(200, ws.max_row))
    areas = []
    for r in range(data_row, last_row + 1):
        desc = ws.cell(row=r, column=5).value
        if not desc:
            continue
        validade_laudo = as_date(ws.cell(row=r, column=20).value)
        validade_is = as_date(ws.cell(row=r, column=23).value)
        areas.append({
            "codigo_ld": clean(ws.cell(row=r, column=2).value),
            "descricao": str(desc).strip(),
            "status": clean(ws.cell(row=r, column=26).value),
            "adequacao_geral": ws.cell(row=r, column=16).value if isinstance(ws.cell(row=r, column=16).value, (int, float)) else None,
            "validade_laudo": validade_laudo.isoformat() if validade_laudo else None,
            "validade_is": validade_is.isoformat() if validade_is else None,
            "dossie": clean(ws.cell(row=r, column=27).value),
            "pendencia": clean(ws.cell(row=r, column=29).value),
        })
    return areas


def read_nao_conformidades(wb):
    ws = wb["RIs"]
    data_row = 4
    last_row = last_row_with_value(ws, 1, data_row, max(99, ws.max_row))
    NC_DESC_COL, NC_SEV_COL, NC_STATUS_COL, NC_RESP_COL, NC_DATE_COL = 13, 14, 15, 16, 17
    rows = []
    for r in range(data_row, last_row + 1):
        desc = ws.cell(row=r, column=NC_DESC_COL).value
        if not desc or not str(desc).strip():
            continue
        data_nc = as_date(ws.cell(row=r, column=NC_DATE_COL).value)
        rows.append({
            "area_texto": clean(ws.cell(row=r, column=3).value),
            "numero_ri": clean(ws.cell(row=r, column=1).value),
            "descricao": str(desc).strip(),
            "severidade": ws.cell(row=r, column=NC_SEV_COL).value or "Não informado",
            "status": ws.cell(row=r, column=NC_STATUS_COL).value or "Não informado",
            "responsavel": clean(ws.cell(row=r, column=NC_RESP_COL).value),
            "data": data_nc.isoformat() if data_nc else None,
        })
    return rows


# tipo -> (nome da aba, linha do cabecalho, primeira linha de dado, teto de linhas,
#          colunas de titulo a concatenar, coluna de revisao, coluna de data de emissao,
#          coluna numero MSI, coluna numero J.Mendes, coluna de observacao/ART)
DOC_SHEETS = {
    "Desenhos (DE)": ("Desenhos", 5, 92, [4, 5, 6], 8, 9, 10, 11, 2),
    "Memoriais Descritivos (MD)": ("MDs", 4, 87, [4, 5, 6, 7, 8], 9, 10, 11, 12, 2),
    "Listas de Materiais (LM)": ("LMs", 4, 89, [4, 5, 6, 7, 8], 9, 10, 11, 12, 2),
    "Análises de Risco (MC)": ("MCs", 4, 98, [4, 5, 6, 7, 8], 9, 10, 11, 12, 2),
    "Relatórios de Inspeção (RI)": ("RIs", 4, 99, [4, 5, 6, 7, 8], 9, 10, 11, 12, 2),
}


def read_documentos(wb):
    docs = []
    for tipo, (sheet_name, data_row, floor_last, titulo_cols, rev_col, data_col, msi_col, jm_col, obs_col) in DOC_SHEETS.items():
        ws = wb[sheet_name]
        last_row = last_row_with_value(ws, 1, data_row, max(floor_last, ws.max_row))
        for r in range(data_row, last_row + 1):
            numero = ws.cell(row=r, column=1).value
            if not numero:
                continue
            titulo_parts = [ws.cell(row=r, column=c).value for c in titulo_cols]
            titulo = " — ".join(str(p).strip() for p in titulo_parts if p and str(p).strip())
            data_emissao = as_date(ws.cell(row=r, column=data_col).value)
            revisao = ws.cell(row=r, column=rev_col).value
            docs.append({
                "tipo": tipo,
                "numero": str(numero).strip(),
                "area_texto": clean(ws.cell(row=r, column=3).value),
                "titulo": titulo or None,
                "revisao": revisao if isinstance(revisao, (int, float)) else None,
                "data_emissao": data_emissao.isoformat() if data_emissao else None,
                "numero_msi": clean(ws.cell(row=r, column=msi_col).value),
                "numero_jmendes": clean(ws.cell(row=r, column=jm_col).value),
                "observacao": clean(ws.cell(row=r, column=obs_col).value),
            })
    return docs


def norm(s):
    return (s or "").strip().casefold()


def run_import(config, session):
    api_base_url = config["api_base_url"]

    print(f"Conectando em {api_base_url} como {config['admin_email']}...")
    login(session, api_base_url, config["admin_email"], config["admin_password"])

    existing = api_get(session, api_base_url, "/api/areas")
    if existing and not config["force"]:
        sys.exit(
            f"Já existem {len(existing)} área(s) cadastradas no banco. Rodar de novo duplicaria dados.\n"
            "Se é intencional, rode de novo com --force."
        )

    print(f"Lendo planilha: {config['xlsx_path']}")
    wb = openpyxl.load_workbook(config["xlsx_path"], data_only=True)
    areas = read_areas(wb)
    nao_conformidades = read_nao_conformidades(wb)
    documentos = read_documentos(wb)
    print(f"  áreas: {len(areas)}")
    print(f"  não conformidades: {len(nao_conformidades)}")
    print(f"  documentos: {len(documentos)}")

    print("Importando áreas...")
    area_id_by_desc = {}
    for a in areas:
        created = api_post(session, api_base_url, "/api/areas", a)
        area_id_by_desc[norm(created["descricao"])] = created["id"]
    print(f"  {len(area_id_by_desc)} áreas criadas.")

    print("Importando não conformidades...")
    unmatched_nc = 0
    for n in nao_conformidades:
        area_id = area_id_by_desc.get(norm(n["area_texto"]))
        if area_id is None:
            unmatched_nc += 1
        n["area_id"] = area_id
        api_post(session, api_base_url, "/api/nao-conformidades", n)
    print(f"  {len(nao_conformidades)} não conformidades criadas"
          f" ({unmatched_nc} sem área correspondente exata -- ajuste manual depois pela UI).")

    print("Importando documentos...")
    unmatched_doc = 0
    for d in documentos:
        area_id = area_id_by_desc.get(norm(d["area_texto"]))
        if area_id is None:
            unmatched_doc += 1
        d["area_id"] = area_id
        api_post(session, api_base_url, "/api/documentos", d)
    print(f"  {len(documentos)} documentos criados"
          f" ({unmatched_doc} sem área correspondente exata -- ajuste manual depois pela UI).")

    print("\nImportação concluída. Confira as contagens acima contra a planilha antes de considerá-la aposentada.")


def main():
    load_dotenv(os.path.join(ROOT, ".env"))
    try:
        config = resolve_config(sys.argv, os.environ)
    except ConfigError as err:
        sys.exit(str(err))
    run_import(config, requests.Session())


if __name__ == "__main__":
    main()
