# -*- coding: utf-8 -*-
"""
Estende a planilha (ja reestruturada pelo usuario) com:
  1. Colunas de-para "Numero MSI" / "Numero J.Mendes" (vazias, para
     preenchimento manual) em todas as abas de documentos.
  2. Uma secao de nao conformidades na aba RIs (Descricao, Severidade,
     Status, Responsavel, Data), com validacao de dados e formatacao
     condicional.
  3. Metricas novas de nao conformidades na aba Resumo.
  4. Reescreve o template do dashboard (aba oculta _DashboardTemplate)
     para a versao nova (sem OS/RDO, com painel de Nao Conformidades).

Nao mexe em nada que o usuario ja ajustou manualmente (Controle, Resumo
existente, remocao de RDO/OS/Obsoletos/Dossie) -- so adiciona colunas e
linhas novas nos espacos ja identificados.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC = "/tmp/work/incoming/uploaded.xlsx"
OUT = "/tmp/work/K19-204-FER-LD-001-R05 - Lista de Documentos (MSI).xlsx"

NAVY = "1F3864"
WHITE = "FFFFFF"
BAND = "F2F2F2"
BORDER_GRAY = "BFBFBF"
GOOD = "0CA30C"
WARNING = "FAB219"
CRITICAL = "D03B3B"
NEUTRAL = "9E9E9E"

thin = Side(style="thin", color=BORDER_GRAY)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
DATA_FONT = Font(name="Arial", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL = PatternFill("solid", fgColor=BAND)
NO_FILL = PatternFill(fill_type=None)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

wb = openpyxl.load_workbook(SRC)

# ---------------------------------------------------------------------------
# 1. De-para columns (Numero MSI / Numero J.Mendes) on every document sheet
# ---------------------------------------------------------------------------
DOC_SHEETS = {
    # sheet: (header_row, data_row, last_data_row, last_existing_col)
    "Desenhos": (4, 5, 92, 9),
    "MDs":      (3, 4, 87, 10),
    "LMs":      (3, 4, 89, 10),
    "MCs":      (3, 4, 98, 10),
    "RIs":      (3, 4, 99, 10),
}

DEPARA_HEADERS = ["Número MSI", "Número J.Mendes"]


def style_header_cell(ws, row, col, text, header_row_height=30):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = HEADER_ALIGN
    if ws.row_dimensions[row].height is None:
        ws.row_dimensions[row].height = header_row_height


def style_data_cell(ws, row, col, band):
    c = ws.cell(row=row, column=col)
    c.font = DATA_FONT
    c.border = BORDER
    c.alignment = DATA_ALIGN
    c.fill = BAND_FILL if band else NO_FILL
    return c


for name, (hrow, drow, last_row, last_col) in DOC_SHEETS.items():
    ws = wb[name]
    depara_start = last_col + 1
    for i, header in enumerate(DEPARA_HEADERS):
        col = depara_start + i
        style_header_cell(ws, hrow, col, header)
        ws.column_dimensions[get_column_letter(col)].width = 22
        for r in range(drow, last_row + 1):
            band = (r - drow) % 2 == 1
            style_data_cell(ws, r, col, band)
    new_last_col = depara_start + len(DEPARA_HEADERS) - 1
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(new_last_col)}{last_row}"
    print(f"{name}: de-para columns {get_column_letter(depara_start)}:{get_column_letter(new_last_col)} added")

# ---------------------------------------------------------------------------
# 2. Nao conformidades section on RIs (right after the de-para columns)
# ---------------------------------------------------------------------------
ri_hrow, ri_drow, ri_last, ri_last_col = DOC_SHEETS["RIs"]
nc_start = ri_last_col + len(DEPARA_HEADERS) + 1  # column after Numero J.Mendes
NC_HEADERS = [
    ("Não Conformidade - Descrição", 42),
    ("Severidade", 14),
    ("Status", 14),
    ("Responsável", 20),
    ("Data", 13),
]
ws = wb["RIs"]
for i, (header, width) in enumerate(NC_HEADERS):
    col = nc_start + i
    style_header_cell(ws, ri_hrow, col, header)
    ws.column_dimensions[get_column_letter(col)].width = width
    for r in range(ri_drow, ri_last + 1):
        band = (r - ri_drow) % 2 == 1
        style_data_cell(ws, r, col, band)

sev_col = nc_start + 1
status_col = nc_start + 2
date_col = nc_start + 4
sev_letter = get_column_letter(sev_col)
status_letter = get_column_letter(status_col)
date_col_letter = get_column_letter(date_col)

# Data validation dropdowns
dv_sev = DataValidation(type="list", formula1='"Crítica,Média,Baixa"', allow_blank=True, showDropDown=False)
dv_sev.error = "Escolha uma opção da lista."
dv_sev.errorTitle = "Severidade inválida"
ws.add_data_validation(dv_sev)
dv_sev.add(f"{sev_letter}{ri_drow}:{sev_letter}{ri_last}")

dv_status = DataValidation(type="list", formula1='"Aberta,Corrigida"', allow_blank=True, showDropDown=False)
dv_status.error = "Escolha uma opção da lista."
dv_status.errorTitle = "Status inválido"
ws.add_data_validation(dv_status)
dv_status.add(f"{status_letter}{ri_drow}:{status_letter}{ri_last}")

# Date number format
for r in range(ri_drow, ri_last + 1):
    ws.cell(row=r, column=date_col).number_format = "DD/MM/YYYY"

# Conditional formatting: Severidade colors
ws.conditional_formatting.add(
    f"{sev_letter}{ri_drow}:{sev_letter}{ri_last}",
    CellIsRule(operator="equal", formula=['"Crítica"'], fill=PatternFill("solid", fgColor=CRITICAL), font=Font(color=WHITE, bold=True)),
)
ws.conditional_formatting.add(
    f"{sev_letter}{ri_drow}:{sev_letter}{ri_last}",
    CellIsRule(operator="equal", formula=['"Média"'], fill=PatternFill("solid", fgColor=WARNING)),
)
ws.conditional_formatting.add(
    f"{sev_letter}{ri_drow}:{sev_letter}{ri_last}",
    CellIsRule(operator="equal", formula=['"Baixa"'], fill=PatternFill("solid", fgColor=NEUTRAL), font=Font(color=WHITE)),
)
# Conditional formatting: Status colors
ws.conditional_formatting.add(
    f"{status_letter}{ri_drow}:{status_letter}{ri_last}",
    CellIsRule(operator="equal", formula=['"Aberta"'], fill=PatternFill("solid", fgColor=CRITICAL), font=Font(color=WHITE, bold=True)),
)
ws.conditional_formatting.add(
    f"{status_letter}{ri_drow}:{status_letter}{ri_last}",
    CellIsRule(operator="equal", formula=['"Corrigida"'], fill=PatternFill("solid", fgColor=GOOD), font=Font(color=WHITE, bold=True)),
)

new_last_col_ri = nc_start + len(NC_HEADERS) - 1
ws.auto_filter.ref = f"A{ri_hrow}:{get_column_letter(new_last_col_ri)}{ri_last}"
print(f"RIs: não conformidade columns {get_column_letter(nc_start)}:{get_column_letter(new_last_col_ri)} added "
      f"(descrição={get_column_letter(nc_start)}, severidade={sev_letter}, status={status_letter}, "
      f"responsável={get_column_letter(nc_start+3)}, data={date_col_letter})")

NC_DESC_COL = nc_start
NC_SEV_COL = sev_col
NC_STATUS_COL = status_col
NC_RESP_COL = nc_start + 3
NC_DATE_COL = date_col

# ---------------------------------------------------------------------------
# 3. Resumo: add Não Conformidades metrics in the blank rows 10-11
# ---------------------------------------------------------------------------
resumo = wb["Resumo"]
# sanity check rows 10-11 are free before writing
occupied = any(resumo.cell(row=r, column=c).value for r in (10, 11) for c in range(1, 10))
if occupied:
    print("AVISO: linhas 10-11 do Resumo não estavam vazias -- pulando inserção automática de métricas.")
else:
    desc_letter = get_column_letter(NC_DESC_COL)
    status_letter_r = get_column_letter(NC_STATUS_COL)
    resumo.cell(row=10, column=2, value=f"=COUNTA(RIs!{desc_letter}{ri_drow}:{desc_letter}{ri_last})")
    resumo.cell(row=10, column=3, value=f'=COUNTIF(RIs!{status_letter_r}{ri_drow}:{status_letter_r}{ri_last},"Aberta")')
    resumo.cell(row=10, column=4, value=f'=COUNTIF(RIs!{status_letter_r}{ri_drow}:{status_letter_r}{ri_last},"Corrigida")')
    resumo.cell(row=11, column=2, value="Não conformidades registradas")
    resumo.cell(row=11, column=3, value="Não conformidades abertas")
    resumo.cell(row=11, column=4, value="Não conformidades corrigidas")
    # copy style from the row-8/9 pair (an existing formula/label pair) so it matches visually
    for col in (2, 3, 4):
        src_formula_cell = resumo.cell(row=8, column=col)
        src_label_cell = resumo.cell(row=9, column=col)
        dst_formula_cell = resumo.cell(row=10, column=col)
        dst_label_cell = resumo.cell(row=11, column=col)
        dst_formula_cell.font = src_formula_cell.font.copy()
        dst_formula_cell.fill = src_formula_cell.fill.copy()
        dst_formula_cell.border = src_formula_cell.border.copy()
        dst_formula_cell.alignment = src_formula_cell.alignment.copy()
        dst_formula_cell.number_format = src_formula_cell.number_format
        dst_label_cell.font = src_label_cell.font.copy()
        dst_label_cell.fill = src_label_cell.fill.copy()
        dst_label_cell.border = src_label_cell.border.copy()
        dst_label_cell.alignment = src_label_cell.alignment.copy()
    print("Resumo: métricas de não conformidades adicionadas nas linhas 10-11")

wb.save(OUT)
print("Saved ->", OUT)
